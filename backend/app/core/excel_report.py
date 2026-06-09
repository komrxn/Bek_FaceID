"""Monthly attendance Excel ("Табель") for accountant Зарина.

V1.1: dropped "опоздал N мин" / "ранний уход N мин" — schedules are too
volatile per employee for those metrics to mean anything. The xlsx now
shows the honest day shape: when came, when left, hours worked, or
"Не отметился" for enrolled employees who didn't mark.

Output structure:
  Sheet 1 — "Табель YYYY-MM"
    Header row:   employee | day 1 | day 2 | … | day N | Часов за месяц
    Each data cell: "П HH:MM\\nУ HH:MM\\nЧасов X.X" or "Не отметился"
                    or "П HH:MM\\nна месте" (came but no went).
    Conditional fill:
      green  — completed (came + went)
      blue   — currently on shift (came, no went)
      gray   — absent / "не отметился"
  Sheet 2 — "Сводка"
    Per employee: Отдел, Должность, Часов всего, Дней отработано,
                  Дней не отметился.

Reuses `app.core.attendance_metrics.derive_day_metrics` so the dashboard
and the Excel always show the same numbers.
"""

from __future__ import annotations

import calendar
from collections.abc import Iterable
from datetime import date
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.core.attendance_metrics import (
    AttendanceEvent as MetricsEvent,
    DayMetrics,
    derive_day_metrics,
    bucket_events_by_shift_day,
)

# ---- Styling ---------------------------------------------------------------

PALETTE = {
    "header_bg":   "1F2937",
    "header_text": "FFFFFF",
    "fill_green":  "DCFCE7",   # completed
    "fill_blue":   "DBEAFE",   # currently on shift
    "fill_gray":   "F1F5F9",   # absent / не отметился
    "text_green":  "166534",
    "text_blue":   "1E40AF",
    "text_gray":   "475569",
}

DEPARTMENT_LABEL_RU = {
    "hall": "Зал",
    "kitchen": "Кухня",
    "other": "Прочий штат",
}

THIN = Side(style="thin", color="E5E7EB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

H_HEADER = Font(name="Calibri", size=11, bold=True, color=PALETTE["header_text"])
H_DAY = Font(name="Calibri", size=10, bold=True)
H_NAME = Font(name="Calibri", size=11, bold=True)
H_CELL = Font(name="Calibri", size=9)


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _tone_for(metrics: DayMetrics) -> str:
    """Return one of 'green', 'blue', 'gray' for the cell color."""
    if not metrics.is_present:
        return "gray"
    if metrics.went_at is None:
        return "blue"
    return "green"


def _cell_text(metrics: DayMetrics) -> str:
    if not metrics.is_present:
        return "Не отметился"
    parts: list[str] = []
    if metrics.came_at is not None:
        parts.append(f"П {metrics.came_at.strftime('%H:%M')}")
    if metrics.went_at is not None:
        parts.append(f"У {metrics.went_at.strftime('%H:%M')}")
        if metrics.worked_hours > 0:
            parts.append(f"{metrics.worked_hours:.1f} ч.")
    else:
        parts.append("на месте")
    return "\n".join(parts)


def _days_in_month(year: int, month: int) -> list[date]:
    _, ndays = calendar.monthrange(year, month)
    return [date(year, month, d) for d in range(1, ndays + 1)]


# ---- Data shapes -----------------------------------------------------------


class EmployeeForReport:
    __slots__ = ("id", "full_name", "position", "department")

    def __init__(
        self, *, id: int, full_name: str, position: str, department: str
    ) -> None:
        self.id = id
        self.full_name = full_name
        self.position = position
        self.department = department


# ---- Builder ---------------------------------------------------------------


def build_xlsx(
    *,
    employees: list[EmployeeForReport],
    events_by_employee: dict[int, Iterable[MetricsEvent]],
    year: int,
    month: int,
    tz_name: str,
    shift_day_cutoff_hour: int = 4,
) -> bytes:
    days = _days_in_month(year, month)

    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = f"Табель {year:04d}-{month:02d}"

    # Frozen header row + employee column.
    ws.freeze_panes = "B2"

    # Header
    ws.cell(row=1, column=1, value="Сотрудник").font = H_HEADER
    ws.cell(row=1, column=1).fill = _fill(PALETTE["header_bg"])
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="left", vertical="center")
    for i, d in enumerate(days, start=2):
        c = ws.cell(row=1, column=i, value=f"{d.day:02d}")
        c.font = H_HEADER
        c.fill = _fill(PALETTE["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    total_col = len(days) + 2
    tcell = ws.cell(row=1, column=total_col, value="Часов за месяц")
    tcell.font = H_HEADER
    tcell.fill = _fill(PALETTE["header_bg"])
    tcell.alignment = Alignment(horizontal="center", vertical="center")

    # Column widths
    ws.column_dimensions[get_column_letter(1)].width = 32
    for i in range(2, total_col):
        ws.column_dimensions[get_column_letter(i)].width = 12
    ws.column_dimensions[get_column_letter(total_col)].width = 16
    ws.row_dimensions[1].height = 24

    # Body
    summary: list[tuple[EmployeeForReport, dict[str, float | int]]] = []
    for row_idx, emp in enumerate(employees, start=2):
        dept_label = DEPARTMENT_LABEL_RU.get(emp.department, emp.department)
        ws.cell(
            row=row_idx,
            column=1,
            value=f"{emp.full_name}\n{dept_label} · {emp.position}",
        )
        name_cell = ws.cell(row=row_idx, column=1)
        name_cell.font = H_NAME
        name_cell.alignment = Alignment(vertical="center", wrap_text=True)
        name_cell.border = BORDER

        events = list(events_by_employee.get(emp.id, []))
        buckets = bucket_events_by_shift_day(
            events, tz_name=tz_name, cutoff_hour=shift_day_cutoff_hour
        )

        total_hours = 0.0
        days_present = 0
        days_unmarked = 0

        for col_offset, d in enumerate(days):
            metrics = derive_day_metrics(
                buckets.get(d, []),
                tz_name=tz_name,
                shift_day_cutoff_hour=shift_day_cutoff_hour,
                target_day=d,
            )
            tone = _tone_for(metrics)
            cell = ws.cell(row=row_idx, column=2 + col_offset, value=_cell_text(metrics))
            cell.font = H_CELL
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.fill = _fill(PALETTE[f"fill_{tone}"])
            cell.border = BORDER

            total_hours += metrics.worked_hours
            if metrics.is_present:
                days_present += 1
            else:
                days_unmarked += 1

        total_cell = ws.cell(row=row_idx, column=total_col, value=round(total_hours, 1))
        total_cell.font = H_NAME
        total_cell.alignment = Alignment(horizontal="center", vertical="center")
        total_cell.border = BORDER

        summary.append(
            (
                emp,
                {
                    "total_hours": round(total_hours, 1),
                    "days_present": days_present,
                    "days_unmarked": days_unmarked,
                    "days_total": len(days),
                },
            )
        )

        ws.row_dimensions[row_idx].height = 36

    # ---- Sheet 2 — summary -------------------------------------------------
    ws2 = wb.create_sheet("Сводка")
    headers = [
        "Сотрудник",
        "Отдел",
        "Должность",
        "Часов всего",
        "Дней отработано",
        "Дней не отметился",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = H_HEADER
        c.fill = _fill(PALETTE["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")

    for row_idx, (emp, s) in enumerate(summary, start=2):
        ws2.cell(row=row_idx, column=1, value=emp.full_name).font = H_NAME
        ws2.cell(
            row=row_idx,
            column=2,
            value=DEPARTMENT_LABEL_RU.get(emp.department, emp.department),
        ).font = H_CELL
        ws2.cell(row=row_idx, column=3, value=emp.position).font = H_CELL
        ws2.cell(row=row_idx, column=4, value=s["total_hours"]).font = H_CELL
        ws2.cell(
            row=row_idx,
            column=5,
            value=f"{s['days_present']} / {s['days_total']}",
        ).font = H_CELL
        unmarked_c = ws2.cell(row=row_idx, column=6, value=s["days_unmarked"])
        unmarked_c.font = H_CELL
        if s["days_unmarked"] > 0:
            unmarked_c.fill = _fill(PALETTE["fill_gray"])

    for col, width in enumerate([32, 14, 24, 14, 18, 18], start=1):
        ws2.column_dimensions[get_column_letter(col)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def build_daily_xlsx(
    *,
    employees: list[EmployeeForReport],
    events_by_employee: dict[int, Iterable[MetricsEvent]],
    day: date,
    tz_name: str,
    shift_day_cutoff_hour: int = 4,
) -> bytes:
    """Single-day report: one row per employee with came/went/hours/status.

    Mirrors the admin dashboard's "Today" table so a manager can hand a
    daily attendance snapshot to anyone who asks. Same tone coding as the
    monthly grid for visual consistency.
    """
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = f"Посещ {day.isoformat()}"
    ws.freeze_panes = "A2"

    headers = [
        "Сотрудник",
        "Отдел",
        "Должность",
        "Пришёл",
        "Ушёл",
        "Часов",
        "Статус",
    ]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = H_HEADER
        c.fill = _fill(PALETTE["header_bg"])
        c.alignment = Alignment(horizontal="center", vertical="center")

    totals = {"working_now": 0, "completed": 0, "absent": 0}
    for row_idx, emp in enumerate(employees, start=2):
        events = list(events_by_employee.get(emp.id, []))
        metrics = derive_day_metrics(
            events,
            tz_name=tz_name,
            shift_day_cutoff_hour=shift_day_cutoff_hour,
            target_day=day,
        )
        tone = _tone_for(metrics)
        if metrics.is_present and metrics.went_at is None:
            totals["working_now"] += 1
            status = "На смене"
        elif metrics.is_present:
            totals["completed"] += 1
            status = "Отработал"
        else:
            totals["absent"] += 1
            status = "Не отметился"

        ws.cell(row=row_idx, column=1, value=emp.full_name).font = H_NAME
        ws.cell(
            row=row_idx, column=2,
            value=DEPARTMENT_LABEL_RU.get(emp.department, emp.department),
        ).font = H_CELL
        ws.cell(row=row_idx, column=3, value=emp.position).font = H_CELL
        ws.cell(
            row=row_idx, column=4,
            value=metrics.came_at.strftime("%H:%M") if metrics.came_at else "—",
        ).font = H_CELL
        ws.cell(
            row=row_idx, column=5,
            value=metrics.went_at.strftime("%H:%M") if metrics.went_at else "—",
        ).font = H_CELL
        ws.cell(
            row=row_idx, column=6,
            value=round(metrics.worked_hours, 1) if metrics.worked_hours > 0 else "—",
        ).font = H_CELL
        sc = ws.cell(row=row_idx, column=7, value=status)
        sc.font = H_CELL
        sc.fill = _fill(PALETTE[f"fill_{tone}"])
        sc.alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, 8):
            ws.cell(row=row_idx, column=col).border = BORDER
            if col in (4, 5, 6):  # numeric/time cols centered
                ws.cell(row=row_idx, column=col).alignment = Alignment(
                    horizontal="center", vertical="center"
                )

    # Totals row
    summary_row = len(employees) + 3
    totals_label = (
        f"Итого за день: на смене {totals['working_now']}, "
        f"отработали {totals['completed']}, не отметились {totals['absent']}"
    )
    tc = ws.cell(row=summary_row, column=1, value=totals_label)
    tc.font = H_NAME
    ws.merge_cells(start_row=summary_row, start_column=1, end_row=summary_row, end_column=7)
    tc.alignment = Alignment(horizontal="left", vertical="center")

    for col, width in enumerate([32, 14, 24, 10, 10, 10, 18], start=1):
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[1].height = 24

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
